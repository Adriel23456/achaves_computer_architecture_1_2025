from pathlib import Path
import openpyxl
from openpyxl import Workbook
import threading
import queue
from enum import Enum
from typing import Union, Tuple, Optional, Any, List
import time
import os
from datetime import datetime


class ActionType(Enum):
    """Tipos de acciones que puede realizar TableControl"""
    READ = "READ"
    WRITE = "WRITE"


class DataType(Enum):
    """Tipos de datos soportados"""
    STRING = "string"
    INT = "int"
    BINARY = "binary"
    HEX = "hex"


class TableAction:
    """Representa una acción en la cola"""
    def __init__(self, action_type: ActionType, row: int, column: int, 
                 content: Any = None, callback: callable = None):
        self.action_type = action_type
        self.row = row
        self.column = column
        self.content = content
        self.callback = callback
        self.timestamp = datetime.now()


class TableControl:
    """
    Clase Singleton para controlar lectura/escritura de archivos Excel.
    Implementa una cola FIFO para procesar acciones de forma manual.
    """
    _instance = None
    _lock = threading.Lock()
    
    def __new__(cls, *args, **kwargs):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super(TableControl, cls).__new__(cls)
                    cls._instance._initialized = False
        return cls._instance
    
    def __init__(self, filename: str = None):
        """
        Inicializa TableControl con un archivo Excel.
        
        Args:
            filename: Nombre del archivo Excel a utilizar
        """
        if self._initialized:
            return
        
        # Usar ruta absoluta basada en la ubicación de este archivo
        current_dir = Path(os.path.dirname(os.path.abspath(__file__)))
        parent_dir = current_dir.parent  # Subir un nivel desde ExtraPrograms
        assets_dir = parent_dir / "assets"

        # Si se pasa un filename relativo, convertirlo a absoluto
        if filename is None or filename == "assets/table_data.xlsx":
            self.filename = str(assets_dir / "table_data.xlsx")
        else:
            # Si es una ruta relativa, hacerla absoluta respecto al parent_dir
            if not os.path.isabs(filename):
                self.filename = str(parent_dir / filename)
            else:
                self.filename = filename
            
        self.action_queue = queue.Queue()
        self.workbook = None
        self.worksheet = None
        
        # Crear directorio assets si no existe
        os.makedirs(assets_dir, exist_ok=True)
        
        # Inicializar archivo Excel
        self._initialize_excel()
        
        self._initialized = True
        
    def _initialize_excel(self):
        """Inicializa o carga el archivo Excel"""
        try:
            if os.path.exists(self.filename):
                self.workbook = openpyxl.load_workbook(self.filename)
                self.worksheet = self.workbook.active
            else:
                self.workbook = Workbook()
                self.worksheet = self.workbook.active
                self.worksheet.title = "Datos"
                
                # Ajustar ancho de columnas
                for column in self.worksheet.columns:
                    self.worksheet.column_dimensions[column[0].column_letter].width = 25
                
                self._save_workbook()
                print(f"✓ Archivo '{self.filename}' creado exitosamente")
                
        except Exception as e:
            print(f"✗ Error al inicializar Excel: {e}")
            raise
    
    def _save_workbook(self):
        """Guarda el archivo Excel de forma segura"""
        try:
            # Usar la ruta absoluta que ya está en self.filename
            self.workbook.save(self.filename)
        except PermissionError:
            # Si el archivo está abierto, intentar guardar con otro nombre
            current_dir = Path(os.path.dirname(os.path.abspath(__file__)))
            parent_dir = current_dir.parent
            backup_name = str(parent_dir / "assets" / "table_data_temp.xlsx")
            self.workbook.save(backup_name)
            print(f"⚠ Archivo principal bloqueado, guardado en: {backup_name}")
        except Exception as e:
            print(f"✗ Error al guardar: {e}")
    
    def read_immediate(self, row: int, column: int) -> Tuple[DataType, Any]:
        """
        Lee inmediatamente un valor de la celda especificada, sin usar la cola.
        Retorna una tupla con (DataType, valor_leido).
        
        Args:
            row: Fila de donde leer (1-indexed)
            column: Columna de donde leer (1-indexed)
            
        Returns:
            Tuple[DataType, Any]: Tupla con el tipo de dato detectado y el valor
            
        Raises:
            Exception: Si hay error al leer la celda
        """
        try:
            # Recargar el archivo para asegurar datos actualizados
            self._reload_workbook()
            
            # Leer valor de la celda
            cell_value = self.worksheet.cell(row=row, column=column).value
            
            # Detectar tipo y procesar valor
            data_type, processed_value = self._detect_data_type(cell_value)
            
            return data_type, processed_value
            
        except Exception as e:
            print(f"✗ Error en lectura inmediata: {e}")
            raise
    
    def _reload_workbook(self):
        """Recarga el archivo Excel para obtener datos actualizados"""
        try:
            if os.path.exists(self.filename):
                # Cerrar workbook actual si existe
                if self.workbook:
                    self.workbook.close()
                
                # Recargar archivo
                self.workbook = openpyxl.load_workbook(self.filename)
                self.worksheet = self.workbook.active
            else:
                raise FileNotFoundError(f"Archivo no encontrado: {self.filename}")
        except Exception as e:
            print(f"⚠ Error al recargar archivo: {e}")
            raise
    
    def _uint32_to_int32(self, value: int) -> int:
        """Convierte un valor uint32 a int32 con signo usando complemento a 2"""
        value = value & 0xFFFFFFFF  # Asegurar que es de 32 bits
        if value > 2147483647:
            return value - 4294967296
        return value

    def _detect_data_type(self, cell_value: Any) -> Tuple[DataType, Any]:
        """
        Detecta el tipo de dato basado en el contenido de la celda.
        
        Args:
            cell_value: Valor leído de la celda
            
        Returns:
            Tuple[DataType, Any]: Tupla con el tipo detectado y valor procesado
        """
        if cell_value is None:
            return DataType.STRING, None
        
        # Convertir a string para análisis
        value_str = str(cell_value).strip()
        
        # Detectar patrones de formato de datos
        if value_str.startswith("String: '") and value_str.endswith("'"):
            # Formato: String: 'texto'
            actual_value = value_str[9:-1]  # Remover "String: '" y "'"
            return DataType.STRING, actual_value
            
        elif value_str.startswith("Int: 0d"):
            # Formato: Int: 0d12345
            try:
                decimal_part = value_str[7:]  # Remover "Int: 0d"
                actual_value = int(decimal_part)
                # Aplicar conversión uint32 a int32
                return DataType.INT, self._uint32_to_int32(actual_value)
            except ValueError:
                return DataType.STRING, value_str
                
        elif value_str.startswith("Binary: 0b") or value_str.startswith("Binary(large): 0b"):
            # Formato: Binary: 0b101010 o Binary(large): 0b101010
            if "Binary(large):" in value_str:
                binary_part = value_str[value_str.find("0b") + 2:]
                return DataType.BINARY, f"0b{binary_part}"
            else:
                binary_part = value_str[11:]  # Remover "Binary: 0b"
                try:
                    # Intentar convertir a decimal si no es muy grande
                    if len(binary_part) <= 64:
                        actual_value = int(f"0b{binary_part}", 2)
                        # Aplicar conversión uint32 a int32
                        return DataType.BINARY, self._uint32_to_int32(actual_value)
                    else:
                        return DataType.BINARY, f"0b{binary_part}"
                except ValueError:
                    return DataType.STRING, value_str
                    
        elif value_str.startswith("Hex: 0x") or value_str.startswith("Hex(large): 0x"):
            # Formato: Hex: 0xDEADBEEF o Hex(large): 0xDEADBEEF
            if "Hex(large):" in value_str:
                hex_part = value_str[value_str.find("0x") + 2:]
                return DataType.HEX, f"0x{hex_part}"
            else:
                hex_part = value_str[7:]  # Remover "Hex: 0x"
                try:
                    # Intentar convertir a decimal si no es muy grande
                    if len(hex_part) <= 16:
                        actual_value = int(f"0x{hex_part}", 16)
                        # Aplicar conversión uint32 a int32
                        return DataType.HEX, self._uint32_to_int32(actual_value)
                    else:
                        return DataType.HEX, f"0x{hex_part}"
                except ValueError:
                    return DataType.STRING, value_str
        
        # Si no coincide con ningún patrón, intentar detectar tipo nativo
        elif isinstance(cell_value, int) and not isinstance(cell_value, bool):
            # Aplicar conversión de uint32 a int32 con signo
            return DataType.INT, self._uint32_to_int32(cell_value)
        
        else:
            # Cualquier otro caso es string
            return DataType.STRING, value_str
    
    def read_immediate_as_string(self, row: int, column: int) -> str:
        """
        Lee inmediatamente un valor como string, útil para debug.
        
        Args:
            row: Fila de donde leer (1-indexed)
            column: Columna de donde leer (1-indexed)
            
        Returns:
            str: Valor de la celda como string
        """
        try:
            self._reload_workbook()
            cell_value = self.worksheet.cell(row=row, column=column).value
            result = str(cell_value) if cell_value is not None else "None"
            print(f"📄 Lectura como string: '{result}' desde [{row}, {column}]")
            return result
        except Exception as e:
            print(f"✗ Error en lectura como string: {e}")
            raise
    
    def execute_all(self):
        """Ejecuta todas las acciones pendientes en la cola en orden FIFO"""
        executed = 0
        total_actions = self.action_queue.qsize()
        
        if total_actions == 0:
            print("⚠ No hay acciones pendientes en la cola")
            return 0
        
        while not self.action_queue.empty():
            try:
                action = self.action_queue.get_nowait()
                self._execute_action(action)
                self.action_queue.task_done()
                executed += 1
            except queue.Empty:
                break
            except Exception as e:
                print(f"✗ Error ejecutando acción: {e}")
        return executed
    
    def execute_reads_only(self):
        """Ejecuta solo las acciones de lectura en orden FIFO"""
        # Obtener todas las acciones de la cola
        actions = []
        while not self.action_queue.empty():
            try:
                actions.append(self.action_queue.get_nowait())
            except queue.Empty:
                break
        
        # Separar lecturas y escrituras
        reads = [a for a in actions if a.action_type == ActionType.READ]
        writes = [a for a in actions if a.action_type == ActionType.WRITE]
        
        if len(reads) == 0:
            print("⚠ No hay acciones de lectura en la cola")
            # Devolver las escrituras a la cola
            for action in writes:
                self.action_queue.put(action)
            return 0
        
        print(f"\n📖 Ejecutando {len(reads)} acciones de LECTURA...")
        
        # Ejecutar solo las lecturas
        executed = 0
        for action in reads:
            try:
                self._execute_action(action)
                executed += 1
            except Exception as e:
                print(f"✗ Error ejecutando lectura: {e}")
        
        # Devolver las escrituras a la cola
        for action in writes:
            self.action_queue.put(action)
        
        print(f"✓ {executed} lecturas ejecutadas")
        print(f"💾 {len(writes)} escrituras permanecen en la cola")
        return executed
    
    def execute_writes_only(self):
        """Ejecuta solo las acciones de escritura en orden FIFO"""
        # Obtener todas las acciones de la cola
        actions = []
        while not self.action_queue.empty():
            try:
                actions.append(self.action_queue.get_nowait())
            except queue.Empty:
                break
        
        # Separar lecturas y escrituras
        reads = [a for a in actions if a.action_type == ActionType.READ]
        writes = [a for a in actions if a.action_type == ActionType.WRITE]
        
        if len(writes) == 0:
            print("⚠ No hay acciones de escritura en la cola")
            # Devolver las lecturas a la cola
            for action in reads:
                self.action_queue.put(action)
            return 0
        
        print(f"\n✏️ Ejecutando {len(writes)} acciones de ESCRITURA...")
        
        # Ejecutar solo las escrituras
        executed = 0
        for action in writes:
            try:
                self._execute_action(action)
                executed += 1
            except Exception as e:
                print(f"✗ Error ejecutando escritura: {e}")
        
        # Devolver las lecturas a la cola
        for action in reads:
            self.action_queue.put(action)
        
        print(f"✓ {executed} escrituras ejecutadas")
        print(f"📖 {len(reads)} lecturas permanecen en la cola")
        return executed
    
    def reorder_reads_first(self):
        """Reordena la cola para que todas las lecturas estén primero, luego las escrituras"""
        # Obtener todas las acciones de la cola
        actions = []
        while not self.action_queue.empty():
            try:
                actions.append(self.action_queue.get_nowait())
            except queue.Empty:
                break
        
        # Separar lecturas y escrituras manteniendo su orden original
        reads = [a for a in actions if a.action_type == ActionType.READ]
        writes = [a for a in actions if a.action_type == ActionType.WRITE]
        
        # Volver a llenar la cola con lecturas primero, luego escrituras
        for action in reads:
            self.action_queue.put(action)
        for action in writes:
            self.action_queue.put(action)
        
        print(f"\n🔄 Cola reordenada: {len(reads)} lecturas primero, {len(writes)} escrituras después")
        print(f"📊 Total de acciones en cola: {self.get_queue_size()}")
    
    def _execute_action(self, action: TableAction):
        """Ejecuta una acción individual"""
        
        if action.action_type == ActionType.WRITE:
            self._execute_write(action)
        elif action.action_type == ActionType.READ:
            self._execute_read(action)
    
    def _execute_write(self, action: TableAction):
        """Ejecuta una acción de escritura"""
        try:
            # Procesar y validar el contenido
            content, display_value = self._process_content(action.content)
            
            # Escribir en la celda
            self.worksheet.cell(row=action.row, column=action.column, value=display_value)
            
            # Guardar archivo
            self._save_workbook()
            
            # Ejecutar callback si existe
            if action.callback:
                action.callback(True, content)
                
        except Exception as e:
            print(f"✗ Error en escritura: {e}")
            if action.callback:
                action.callback(False, str(e))
    
    def _execute_read(self, action: TableAction):
        """Ejecuta una acción de lectura"""
        try:
            # Leer valor de la celda
            cell_value = self.worksheet.cell(row=action.row, column=action.column).value
            
            print(f"✓ Leído '{cell_value}' de celda [{action.row}, {action.column}]")
            
            # Ejecutar callback con el valor
            if action.callback:
                action.callback(True, cell_value)
            
            return cell_value
            
        except Exception as e:
            print(f"✗ Error en lectura: {e}")
            if action.callback:
                action.callback(False, str(e))
            return None
    
    def _process_content(self, content: Any) -> Tuple[Any, str]:
        """
        Procesa el contenido según su tipo y retorna (valor_procesado, valor_display)
        
        Soporta:
        - String: 'texto' o "texto"
        - Int: 0d12345 (decimal, soporta negativos con complemento a 2)
        - Binary: 0b1010101
        - Hex: 0xDEADBEEF
        """
        if content is None:
            return None, "None"
        
        content_str = str(content)
        
        # String con comillas
        if ((content_str.startswith("'") and content_str.endswith("'")) or 
            (content_str.startswith('"') and content_str.endswith('"'))):
            # Remover comillas
            actual_content = content_str[1:-1]
            return actual_content, f"String: '{actual_content}'"
        
        # Entero decimal con prefijo 0d
        elif content_str.startswith("0d"):
            try:
                decimal_value = int(content_str[2:])
                # Validar rango de int32
                if decimal_value < -2147483648 or decimal_value > 2147483647:
                    raise ValueError(f"Valor {decimal_value} fuera de rango para int32")
                # Almacenar el valor tal cual (negativo o positivo)
                return decimal_value, f"Int: 0d{decimal_value}"
            except ValueError as e:
                raise ValueError(f"Valor decimal inválido: {content_str} - {e}")
        
        # Binario con prefijo 0b
        elif content_str.startswith("0b"):
            try:
                binary_str = content_str[2:]
                # Validar que solo contenga 0 y 1
                if all(c in '01' for c in binary_str):
                    # Para números grandes, guardar como string
                    if len(binary_str) > 64:
                        return binary_str, f"Binary(large): 0b{binary_str}"
                    else:
                        decimal_value = int(content_str, 2)
                        return decimal_value, f"Binary: 0b{binary_str}"
                else:
                    raise ValueError(f"Valor binario inválido: {content_str}")
            except ValueError as e:
                raise ValueError(f"Error procesando binario: {e}")
        
        # Hexadecimal con prefijo 0x o 0X
        elif content_str.lower().startswith("0x"):
            try:
                hex_str = content_str[2:]
                # Validar caracteres hexadecimales
                if all(c in '0123456789abcdefABCDEF' for c in hex_str):
                    # Para números grandes, guardar como string
                    if len(hex_str) > 16:
                        return hex_str.upper(), f"Hex(large): 0x{hex_str.upper()}"
                    else:
                        decimal_value = int(content_str, 16)
                        return decimal_value, f"Hex: 0x{hex_str.upper()}"
                else:
                    raise ValueError(f"Valor hexadecimal inválido: {content_str}")
            except ValueError as e:
                raise ValueError(f"Error procesando hexadecimal: {e}")
        
        # Si es un entero simple sin prefijo
        elif isinstance(content, int) and not isinstance(content, bool):
            # Validar rango y almacenar tal cual
            if content < -2147483648 or content > 2147483647:
                raise ValueError(f"Valor {content} fuera de rango para int32")
            return content, f"Int: 0d{content}"
        
        # Cualquier otro caso, tratar como string sin comillas
        else:
            return content_str, f"String: '{content_str}'"
    
    def write(self, row: int, column: int, content: Any, callback: callable = None):
        """
        Agrega una acción de escritura a la cola.
        
        Args:
            row: Fila donde escribir (1-indexed)
            column: Columna donde escribir (1-indexed)
            content: Contenido a escribir (formatos: 'string', 0d123, 0b101, 0xABC)
            callback: Función a llamar cuando se complete (opcional)
        """
        action = TableAction(ActionType.WRITE, row, column, content, callback)
        self.action_queue.put(action)
    
    def read(self, row: int, column: int, callback: callable = None):
        """
        Agrega una acción de lectura a la cola.
        
        Args:
            row: Fila de donde leer (1-indexed)
            column: Columna de donde leer (1-indexed)
            callback: Función a llamar con el resultado
        """
        action = TableAction(ActionType.READ, row, column, callback=callback)
        self.action_queue.put(action)
        print(f"➕ Acción de lectura agregada a la cola (Total: {self.get_queue_size()})")
    
    def get_queue_size(self) -> int:
        """Retorna el número de acciones pendientes en la cola"""
        return self.action_queue.qsize()
    
    def clear_queue(self):
        """Limpia todas las acciones pendientes en la cola"""
        cleared = 0
        while not self.action_queue.empty():
            try:
                self.action_queue.get_nowait()
                cleared += 1
            except queue.Empty:
                break
        print(f"🗑️ {cleared} acciones eliminadas de la cola")
        return cleared
    
    def __del__(self):
        """Limpieza al destruir la instancia"""
        if hasattr(self, 'workbook') and self.workbook:
            self.workbook.close()